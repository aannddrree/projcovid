import openpyxl
import pymongo
from pathlib import Path

myclient = pymongo.MongoClient("mongodb://localhost:27017/")
mydb = myclient["covidDB"]
mycol = mydb["colcovid"]

#Diretório do arquivo
xlsx_file = Path('C:\\tmp\\arquivo\\covidfull.xlsx')

wb_obj = openpyxl.load_workbook(xlsx_file)
sheet = wb_obj.active

col_names = []

print('Qtd colunas: ' + str(sheet.max_column))

for column in sheet.iter_cols(1, sheet.max_column):
    col_names.append(column[0].value)

for row in sheet.iter_rows(1, sheet.max_row):
    mydict = {col_names[0]: row[0].value, col_names[1]: row[1].value, col_names[2]: row[2].value, col_names[3]: row[3].value, col_names[4]: row[4].value,col_names[5]: row[5].value, col_names[6]: row[6].value, col_names[7]: row[7].value, col_names[8]: row[8].value, col_names[9]: row[9].value, col_names[10]: row[10].value,col_names[10]: row[10].value, col_names[11]: row[11].value, col_names[12]: row[12].value, col_names[13]: row[13].value, col_names[14]: row[14].value,col_names[15]: row[15].value, col_names[16]: row[16].value, col_names[17]: row[17].value, col_names[18]: row[18].value, col_names[19]: row[19].value, col_names[20]: row[20].value, col_names[21]: row[21].value}
    mycol.insert_one(mydict)

print('Fim')